import asyncio
import logging
from io import BytesIO
from typing import Any, Dict, Tuple
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from shared.safe_naming import MessageType, TopicNames
from shared.utils import createResponseMessage, getMenuOptions
from shared.safe_naming import UserState

logger = logging.getLogger(__name__)


class QueryProcessor:
    """
    Processes text queries about invoices and generates responses.

    This class handles the receipt, processing, and response for text queries
    about invoice data. It uses GPT to generate SQL queries, executes them against
    the database, and formats the results as downloadable Excel files.

    Attributes:
        kafka_handler: Handler for Kafka messaging operations
        s3_client: Client for S3 operations to store and retrieve files
        db_manager: Database client for executing SQL queries
        gpt_api: Client for GPT API interactions to generate SQL queries
        redis_manager: Client for managing user session states
    """

    def __init__(self, kafka_handler, s3_client, db_manager, gpt_api, redis_manager):
        """
        Initialize the QueryProcessor with necessary clients and handlers.

        Args:
            kafka_handler: Handler for Kafka operations
            s3_client: Client for S3 operations
            db_manager: Database client for executing queries
            gpt_api: Client for GPT API interactions
            redis_manager: Client for managing user session states
        """
        self.kafka_handler = kafka_handler
        self.s3_client = s3_client
        self.db_manager = db_manager
        self.gpt_api = gpt_api
        self.redis_manager = redis_manager
        logger.debug("QueryProcessor initialized")

    async def runService(self):
        """
        Initialize connections and start consuming messages from the topic.

        This method sets up connections to Kafka and Redis, then starts
        listening for messages on the query topic. It runs in an infinite loop
        to maintain the service.

        Returns:
            None
        """
        logger.info("Initializing Kafka connection...")
        await self.kafka_handler._initializeConnection([TopicNames.QUERY_TOPIC.value, TopicNames.RESPONSE_TOPIC.value])
        logger.info("Initializing Redis connection...")
        await self.redis_manager.initializeConnections()
        logger.info("Connection initialized, starting consumer...")
        await self.kafka_handler.consumeFromTopic(TopicNames.QUERY_TOPIC.value, self.processQueryRequest)
        logger.info("Consumer started, entering main loop...")
        while True:
            await asyncio.sleep(1)

    async def processQueryRequest(self, incoming_msg: dict) -> Dict[str, Any]:
        """
        Process a text query request from a user and generate a response.

        This method performs the following steps:
        1. Validates the incoming message content
        2. Uses GPT to generate an SQL query based on the user's text
        3. Executes the query against the database
        4. Formats the results as an Excel file
        5. Uploads the file to S3 and sends the link back to the user

        Args:
            incoming_msg: Dictionary containing message data from WhatsApp
                         including the user's text query

        Returns:
            Dict[str, Any]: Query results if successful, None otherwise
        """
        clients_num = incoming_msg.get("From", None)
        if "Body" not in incoming_msg or not self.validateMessageContent(incoming_msg):
            await self.handleResponse(clients_num, "No Valid message was provided please resend", MessageType.ERROR)
            await self.redis_manager.setSession(clients_num, UserState.AWAITING_TEXT)
            return
        gpt_query = await self.gpt_api.generateInvoiceQuery(incoming_msg["Body"], clients_num)
        if not gpt_query:
            await self.handleResponse(clients_num, "Your request was unclear please try again", MessageType.ERROR)
            await self.redis_manager.setSession(clients_num, UserState.AWAITING_TEXT)
            return

        success, query_results_or_msg = await self.validateExecutedQueryResult(gpt_query, clients_num)
        if not success:
            await self.handleResponse(clients_num, query_results_or_msg, MessageType.ERROR)
            await self.redis_manager.setSession(clients_num, UserState.START)
            return

        outgoing_phone_num, excel_file = await self.convertDataframeToExcel(query_results_or_msg)
        aws_img_path = await self.s3_client.uploadToS3(outgoing_phone_num, excel_file, "xlsx")
        return_file_url = await self.s3_client.generatePresignedUrl(aws_img_path)
        format_response = {
            'body': f'Your excel file is ready for download!\n\n{getMenuOptions()}',
            'to': outgoing_phone_num,
            'media_url': [return_file_url]
        }
        await self.handleResponse(clients_num, format_response)
        await self.redis_manager.setSession(clients_num, UserState.CHOOSING)
        return

    async def convertDictToDataframe(self, dict_to_convert: dict):
        """
        Convert a dictionary of query results to a pandas DataFrame.

        This method takes the database query results and converts them to a
        DataFrame, adding presigned URLs for any raw image URLs.

        Args:
            dict_to_convert: Dictionary containing query results from the database

        Returns:
            pd.DataFrame: DataFrame containing the formatted query results
        """
        query_results_df = pd.DataFrame(dict_to_convert)
        query_results_df['whatsapp_number'] = query_results_df['whatsapp_number'].where(
            query_results_df['whatsapp_number'] == query_results_df['whatsapp_number'].iloc[0], None
        )
        urls = []
        for path in query_results_df['raw_image_url']:
            if path:
                url = await self.s3_client.generatePresignedUrl(path)
                urls.append(url)
            else:
                urls.append(None)
        query_results_df['raw_image_url'] = urls
        logger.debug(f"in utils the df is : {query_results_df}")
        return query_results_df

    async def convertDataframeToExcel(self, queried_df: dict):
        """
        Convert a DataFrame to an Excel file with formatted hyperlinks.

        This method takes a DataFrame of query results and creates an Excel file
        with proper column names and clickable hyperlinks for image URLs.

        Args:
            queried_df: Dictionary containing query results to convert to Excel

        Returns:
            tuple: (phone_number, excel_file_bytes) Tuple containing the user's
                   phone number and the Excel file as bytes
        """
        buffer = BytesIO()
        dataframe = await self.convertDictToDataframe(queried_df)
        dataframe.rename(columns={'raw_image_url': 'Download Link', 'whatsapp_number': 'WhatsApp Number'}, inplace=True)
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            dataframe.to_excel(writer, sheet_name='Sheet1', index=False)
        buffer.seek(0)
        workbook = load_workbook(buffer)
        worksheet = workbook.active
        download_col_idx = dataframe.columns.get_loc('Download Link') + 1
        for row_idx, row in dataframe.iterrows():
            presigned_url = row['Download Link']
            if presigned_url:
                cell = worksheet.cell(row=row_idx + 2, column=download_col_idx)
                cell.value = "Download"
                cell.hyperlink = presigned_url
                cell.font = Font(color="0000FF", underline="single")
        buffer.seek(0)
        workbook.save(buffer)
        buffer.seek(0)
        phone_number = dataframe['WhatsApp Number'].dropna().iloc[0] if not dataframe[
            'WhatsApp Number'].dropna().empty else None
        file_bytes = buffer.getvalue()
        if not file_bytes:
            logger.debug("Excel conversion failed. File is empty.")
            return None, None
        logger.debug(f"Excel file size: {len(file_bytes)} bytes")
        logger.debug(f"The phone number is: {phone_number}")
        return phone_number, file_bytes

    def validateMessageContent(self, dequed_msg: dict):
        """
        Validate that the message contains sufficient content for processing.

        Args:
            dequed_msg: Dictionary containing message data from WhatsApp

        Returns:
            bool: True if the message body has sufficient length, False otherwise
        """
        msg_body = dequed_msg.get("Body", "")
        return True if len(msg_body) >= 20 else False

    async def handleResponse(self, phone_num: str, msg_content: str, msg_type: MessageType = None):
        """
        Send a response message to the user.

        Args:
            phone_num: The phone number to send the message to
            msg_content: The content of the message to send
            msg_type: The type of message (error, success, etc.)

        Returns:
            None.
        """
        response = createResponseMessage(msg_content, phone_num)
        await self.kafka_handler.publishToTopic(TopicNames.RESPONSE_TOPIC.value, response, msg_type)

    async def validateExecutedQueryResult(self, postgres_gpt_query: str, clients_num: str) -> Tuple[bool, Any]:
        """
        Validates and executes a GPT-generated SQL query against the database.

        This method sends the GPT-generated SQL query to the database and handles the following:
        - If the query fails to execute or raises an error, it returns a tuple indicating failure with a relevant error message.
        - If the query executes successfully but returns no records, it returns a tuple indicating failure with a message that no data was found.
        - If the query is successful and returns results, it returns a tuple indicating success along with the query results.

        Args:
            postgres_gpt_query (str): The SQL query string generated by GPT.
            clients_num (str): The WhatsApp number of the client, used for logging and result tracking.

        Returns:
            Tuple[bool, Any]:
                - (False, str) if the query fails or returns no data.
                - (True, list[dict]) if the query succeeds and returns results.
        """
        query_results = await self.db_manager.executeQuery(postgres_gpt_query, clients_num)
        logger.debug(f"Query results from Postgres: {query_results}")

        if query_results is None:
            return False, "An error occurred when trying to retrieve your information. Please try again."

        elif not query_results:
            return False, "No matching information was found for your request."

        else:
            return True, query_results